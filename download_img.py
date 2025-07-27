import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import requests
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry
from io import BytesIO
import os
from urllib.parse import urlparse
from pathlib import Path

# --- 用户配置区 ---
SOURCE_EXCEL_FILES = [
    r'C:\Users\Lenovo\Desktop\BCG\女装-消费品\图片处理-polo.xlsx'
]  # 要处理的文件名列表

URL_COLUMN_NAME = '商品主图'             # 包含图片链接的列名
IMAGE_INSERT_COLUMN_NAME = '图片预览'     # 新增的用于存放图片的列名

# 设置图片在单元格中的尺寸 (可以根据需要调整)
MAX_IMAGE_WIDTH = 150  # 图片最大宽度 (像素)
MAX_IMAGE_HEIGHT = 150 # 图片最大高度 (像素)
# --- 配置区结束 ---

def download_image(url):
    """
    下载图片并返回处理后的Image对象，使用指数退避重试机制

    Args:
        url (str): 图片URL

    Returns:
        Image: 处理后的Image对象，如果下载失败返回None
    """
    # 检查URL是否有效
    if not url or not str(url).startswith('http'):
        print(f"跳过无效或空的URL: {url}")
        return None

    try:
        # 创建带重试机制的session
        session = requests.Session()

        # 配置重试策略：最大重试3次，指数退避因子0.1，对5xx状态码和网络异常重试
        retries = Retry(
            total=3,
            backoff_factor=0.1,
            status_forcelist=[500, 502, 503, 504],
            raise_on_status=False
        )

        # 挂载适配器到session
        adapter = HTTPAdapter(max_retries=retries)
        session.mount('http://', adapter)
        session.mount('https://', adapter)

        # 下载图片
        response = session.get(url, timeout=5)
        response.raise_for_status()  # 如果下载失败，会抛出异常

        # 将图片数据读入内存
        image_data = BytesIO(response.content)
        img = Image(image_data)

        # 计算缩放比例，保持纵横比
        ratio = min(MAX_IMAGE_WIDTH / img.width, MAX_IMAGE_HEIGHT / img.height)
        img.width = img.width * ratio
        img.height = img.height * ratio

        return img

    except requests.exceptions.RequestException as e:
        print(f"下载图片失败: {e}")
        return None
    except Exception as e:
        print(f"处理图片时出错: {e}")
        return None

def process_one_worksheet(ws):
    """
    处理单个工作表中的图片下载和插入

    Args:
        ws: openpyxl的工作表对象
    """
    print(f"\n开始处理工作表: '{ws.title}'")

    # 1. 查找URL列的位置
    url_col_idx = None
    for cell in ws[1]:  # 第一行是标题行
        if cell.value == URL_COLUMN_NAME:
            url_col_idx = cell.column
            break

    # 检查URL列是否存在
    if url_col_idx is None:
        print(f"工作表 '{ws.title}'：找不到名为 '{URL_COLUMN_NAME}' 的列，跳过此工作表。")
        return

    print(f"工作表 '{ws.title}'：找到URL列 '{URL_COLUMN_NAME}' 在第 {url_col_idx} 列。")

    # 2. 在URL列后插入新列
    image_col_idx = url_col_idx + 1
    ws.insert_cols(image_col_idx)

    # 3. 设置新列的标题
    ws.cell(row=1, column=image_col_idx).value = IMAGE_INSERT_COLUMN_NAME
    print(f"工作表 '{ws.title}'：已在第 {url_col_idx} 列后插入新列 '{IMAGE_INSERT_COLUMN_NAME}'。")

    # 4. 调整新列的宽度和行的高度
    image_col_letter = get_column_letter(image_col_idx)
    ws.column_dimensions[image_col_letter].width = MAX_IMAGE_WIDTH / 7  # 粗略转换像素到宽度单位
    for i in range(2, ws.max_row + 1):  # 从第2行开始，因为第1行是标题
        ws.row_dimensions[i].height = MAX_IMAGE_HEIGHT * 0.75  # 粗略转换像素到高度单位
    print(f"工作表 '{ws.title}'：已调整图片列的单元格大小。")

    # 5. 遍历每一行，下载并插入图片
    print(f"工作表 '{ws.title}'：开始下载并插入图片...")
    success_count = 0

    for row_num in range(2, ws.max_row + 1):  # 从第2行开始，因为第1行是标题
        # 获取URL
        url_cell = ws.cell(row=row_num, column=url_col_idx)
        url = url_cell.value

        # 目标单元格（图片插入位置）
        target_cell = ws.cell(row=row_num, column=image_col_idx)

        # 下载图片
        img = download_image(url)

        if img is None:
            print(f"工作表 '{ws.title}'：第 {row_num} 行：跳过无效或下载失败的URL。")
            continue

        # 将图片添加到工作表
        ws.add_image(img, target_cell.coordinate)
        print(f"工作表 '{ws.title}'：第 {row_num} 行：成功插入图片。")
        success_count += 1

    print(f"工作表 '{ws.title}'：处理完成，共成功插入 {success_count} 张图片。")

def process_one_excel_file(source: str, dest: str):
    """
    主函数：读取Excel，对每个工作表下载图片并插入，最后保存新文件。
    """
    # 检查源文件是否存在
    if not os.path.exists(source):
        print(f"错误：源文件 '{source}' 不存在。请检查文件名和路径。")
        return

    # 1. 使用openpyxl读取Excel数据
    try:
        wb = load_workbook(source)
        print(f"成功读取Excel文件，包含 {len(wb.sheetnames)} 个工作表: {wb.sheetnames}")
    except Exception as e:
        print(f"读取Excel文件时出错: {e}")
        return

    # 2. 对每个工作表执行处理
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        process_one_worksheet(ws)

    # 3. 保存最终的Excel文件
    try:
        wb.save(dest)
        print(f"\n所有任务完成！结果已保存到 '{dest}'。")
    except Exception as e:
        print(f"保存最终文件时出错: {e}")

def process_excel_files():
    for src in SOURCE_EXCEL_FILES:
        filename = Path(src)
        dest = filename.parent / f"{filename.stem}-图片已下载.{filename.suffix}"
        process_one_excel_file(src, str(dest))

# --- 运行主函数 ---
if __name__ == "__main__":
    process_excel_files()
